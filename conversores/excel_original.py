"""
Adaptación de ConvExc: todas las hojas de cada original, luego consolidado.

No filtra por selección de hojas. Esa selección es del automatizador,
sobre el consolidado ya convertido.
"""
import os
import tempfile
from pathlib import Path

from openpyxl import Workbook
from python_calamine import CalamineWorkbook

from conversores.config.config_loader import cargar_configuracion
from conversores.reglas import procesar_fila

LIMITE_XLSX = 1_048_576
MAX_DATOS_XLSX = LIMITE_XLSX - 1


def encontrar_encabezado(rows):
    """
    Busca la fila que contiene Fecha y Agente.
    """
    for idx, row in enumerate(rows):
        row_str = [str(x).strip() for x in row]

        if "Fecha" in row_str and "Agente" in row_str:
            return idx

    raise Exception("No se encontró la fila de encabezado")


def _primera_fila(archivo, hoja=None):
    wb = CalamineWorkbook.from_path(str(archivo))
    hojas = list(wb.sheet_names)
    if not hojas:
        return []
    nombre = hoja if hoja in hojas else hojas[0]
    rows = wb.get_sheet_by_name(nombre).to_python()
    if not rows:
        return []
    return [str(x).strip() for x in rows[0]]


def hoja_con_encabezado_en_fila_0(archivo, hoja):
    row0 = _primera_fila(archivo, hoja)
    return "Fecha" in row0 and "Agente" in row0


def parece_ya_convertido(archivo):
    """True si alguna hoja ya trae Fecha/Agente en la fila 0."""
    try:
        wb = CalamineWorkbook.from_path(str(archivo))
        for hoja in wb.sheet_names:
            if hoja_con_encabezado_en_fila_0(archivo, hoja):
                return True
        return False
    except Exception:
        return False


def hojas_listas_para_pandas(archivo, hojas):
    return bool(hojas) and all(
        hoja_con_encabezado_en_fila_0(archivo, hoja) for hoja in hojas
    )


def _crear_xlsx(ruta, encabezado):
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Datos")
    ws.append(encabezado)
    return wb, ws, ruta


def _ruta_temporal(prefijo):
    fd, ruta = tempfile.mkstemp(prefix=prefijo, suffix=".xlsx")
    os.close(fd)
    return ruta


def convertir_excel_original(
    archivo, prefijo="convertido_", hojas_seleccionadas=None
):
    """
    ConvExc sobre las hojas indicadas (o todas, en lote de carpeta).
    El encabezado se busca en las hojas a procesar, no forzado en la hoja 1.
    """
    config = cargar_configuracion()
    archivo = Path(archivo)
    wb_calamine = CalamineWorkbook.from_path(str(archivo))
    hojas_libro = list(wb_calamine.sheet_names)
    if hojas_seleccionadas:
        hojas = [h for h in hojas_libro if h in set(hojas_seleccionadas)]
        if not hojas:
            raise ValueError("Ninguna de las hojas seleccionadas existe en el archivo.")
    else:
        hojas = hojas_libro

    fila_encabezado = None
    hoja_encabezado = None
    encabezado = None
    for nombre_hoja in hojas:
        rows = wb_calamine.get_sheet_by_name(nombre_hoja).to_python()
        try:
            fila_encabezado = encontrar_encabezado(rows)
            hoja_encabezado = nombre_hoja
            encabezado = rows[fila_encabezado]
            break
        except Exception:
            continue

    if encabezado is None:
        raise Exception("No se encontró la fila de encabezado")

    mapa_columnas = {
        str(nombre).strip(): indice
        for indice, nombre in enumerate(encabezado)
        if str(nombre).strip()
    }

    estadisticas = {
        "filas_leidas": 0,
        "filas_eliminadas": 0,
        "filas_modificadas": 0,
        "filas_exportadas": 0,
    }

    ruta_salida = _ruta_temporal(prefijo)
    parte = 1
    filas_actuales = 0
    rutas = []

    wb_out, ws_out, ruta_actual = _crear_xlsx(ruta_salida, encabezado)

    for nombre_hoja in hojas:
        sheet = wb_calamine.get_sheet_by_name(nombre_hoja)
        rows = sheet.to_python()

        if nombre_hoja == hoja_encabezado:
            inicio = fila_encabezado + 2
        else:
            inicio = 0

        for fila in rows[inicio:]:
            fila = list(fila)
            fila = procesar_fila(
                fila,
                config,
                mapa_columnas,
                estadisticas,
            )

            if fila is None:
                continue

            if filas_actuales >= MAX_DATOS_XLSX:
                wb_out.save(ruta_actual)
                rutas.append(ruta_actual)
                parte += 1
                filas_actuales = 0
                ruta_actual = _ruta_temporal(f"{prefijo}p{parte}_")
                wb_out, ws_out, ruta_actual = _crear_xlsx(ruta_actual, encabezado)

            ws_out.append(fila)
            estadisticas["filas_exportadas"] += 1
            filas_actuales += 1

    wb_out.save(ruta_actual)
    rutas.append(ruta_actual)

    print("\nResumen conversión Excel original:")
    print(f"Filas leídas:      {estadisticas['filas_leidas']:,}")
    print(f"Filas eliminadas:  {estadisticas['filas_eliminadas']:,}")
    print(f"Filas modificadas: {estadisticas['filas_modificadas']:,}")
    print(f"Filas exportadas:  {estadisticas['filas_exportadas']:,}")

    if len(rutas) == 1:
        return rutas[0]

    return _consolidar_partes(rutas, prefijo)


def convertir_exceles_originales(
    archivos,
    prefijo="convertido_",
    actualizar_estado=None,
):
    """
    Convierte cada original (todas las hojas) y arma Consolidado.xlsx
    como ConvExc.generar_consolidado: solo la hoja activa de cada parte.
    """
    if not archivos:
        raise ValueError("Debe seleccionar al menos un archivo Excel.")

    convertidos = []
    omitidos = []
    total = len(archivos)

    for indice, archivo in enumerate(archivos, start=1):
        nombre = Path(archivo).name
        mensaje = f"Convirtiendo {indice}/{total}: {nombre}"
        print(f"\n[{indice}/{total}] {nombre}")
        if actualizar_estado:
            progreso = min(90, int(indice / total * 80))
            actualizar_estado(mensaje, progreso)

        try:
            ruta = convertir_excel_original(
                archivo,
                prefijo=f"{prefijo}{indice}_",
            )
            convertidos.append(ruta)
        except Exception as error:
            print(f"ERROR en {nombre}: {error}")
            omitidos.append(f"{nombre} ({error})")

    if not convertidos:
        detalle = "; ".join(omitidos[:8])
        raise ValueError(
            "Ningún Excel se pudo convertir. "
            f"Revisados: {total}. Ejemplos: {detalle}"
        )

    print(
        f"\nArchivos convertidos: {len(convertidos)} / {total}. "
        f"Omitidos: {len(omitidos)}"
    )
    if actualizar_estado:
        actualizar_estado("Consolidando archivos convertidos...", 92)

    consolidado = _consolidar_convertidos(convertidos, prefijo)

    for ruta in convertidos:
        try:
            os.remove(ruta)
        except OSError:
            pass

    return consolidado


def _consolidar_convertidos(rutas, prefijo):
    """Igual que ConvExc.generar_consolidado: wb.active de cada convertido."""
    from openpyxl import load_workbook

    ruta_final = _ruta_temporal(f"{prefijo}cons_")
    wb_final = Workbook(write_only=True)
    ws = wb_final.create_sheet("Datos_1")
    encabezado_escrito = False

    for ruta in rutas:
        wb = load_workbook(ruta, read_only=True, data_only=True)
        hoja = wb.active
        primera_fila = True
        for fila in hoja.iter_rows(values_only=True):
            valores = list(fila)
            if primera_fila:
                primera_fila = False
                if not encabezado_escrito:
                    ws.append(valores)
                    encabezado_escrito = True
                continue
            ws.append(valores)
        wb.close()

    wb_final.save(ruta_final)
    print(f"Consolidado interno: {ruta_final}")
    return ruta_final


def _consolidar_partes(rutas, prefijo):
    """Si un solo original se partió por límite Excel, une las partes como archivos."""
    return _consolidar_convertidos(rutas, prefijo)
